from __future__ import annotations

import csv
import hashlib
import io
import re
from typing import Any
from urllib.parse import urlparse

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import Settings
from app.core.errors import ApiError
from app.integrations.minio import MinioStorage
from app.models import School, SchoolArea, Store
from app.services.stores import attach_image, ensure_categories, replace_categories


HEADERS = ["storeCode", "schoolCode", "areaCode", "name", "category", "address", "imageUrl", "status"]
SCHOOL_HEADERS = ["食堂", "店铺位置", "店铺名称", "店铺图片"]
VALID_STATUSES = {"active", "hidden", "closed"}


def _cell(value: Any) -> str:
    return "" if value is None else str(value).strip()


async def parse_rows(content: bytes, filename: str, settings: Settings, *, target_school_id: int | None = None) -> list[dict[str, str]]:
    if len(content) > settings.max_upload_bytes:
        raise ApiError(413, "FILE_TOO_LARGE", "文件不能超过设定大小")
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix == "csv":
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ApiError(415, "UNSUPPORTED_FILE_TYPE", "CSV 必须使用 UTF-8 编码") from exc
        rows = list(csv.reader(io.StringIO(text)))
    elif suffix == "xlsx":
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
            sheet = workbook.active
            rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
        except Exception as exc:
            raise ApiError(415, "UNSUPPORTED_FILE_TYPE", "XLSX 文件无法解析") from exc
    else:
        raise ApiError(415, "UNSUPPORTED_FILE_TYPE", "仅支持 CSV 或 XLSX 文件")
    if not rows:
        raise ApiError(422, "IMPORT_VALIDATION_FAILED", "导入文件不能为空")
    headers = [_cell(value) for value in rows[0]]
    if target_school_id is not None and headers == SCHOOL_HEADERS:
        data_rows = [row for row in rows[1:] if any(_cell(value) for value in row)]
        if not data_rows:
            raise ApiError(422, "IMPORT_VALIDATION_FAILED", "导入文件没有数据行")
        if len(data_rows) > 1000:
            raise ApiError(422, "IMPORT_VALIDATION_FAILED", "数据行不能超过 1000 行")
        errors: list[dict] = []
        parsed: list[dict[str, str]] = []
        seen: set[tuple[str, str, str]] = set()
        for row_number, row in enumerate(data_rows, start=2):
            values = [_cell(value) for value in row]
            values += [""] * (len(SCHOOL_HEADERS) - len(values))
            item = dict(zip(SCHOOL_HEADERS, values[: len(SCHOOL_HEADERS)]))
            normalized = {
                "areaName": item["食堂"],
                "address": item["店铺位置"],
                "name": item["店铺名称"],
                "imageUrl": item["店铺图片"],
            }
            key = (normalized["areaName"].lower(), normalized["address"].lower(), normalized["name"].lower())
            if key in seen:
                errors.append({"row": row_number, "field": "店铺名称", "code": "DUPLICATE_STORE_IN_FILE", "message": "同一食堂、位置和店铺名称在文件中重复"})
            seen.add(key)
            for field, label, max_length in [
                ("areaName", "食堂", 100),
                ("address", "店铺位置", 255),
                ("name", "店铺名称", 120),
                ("imageUrl", "店铺图片", 2048),
            ]:
                if not normalized[field]:
                    errors.append({"row": row_number, "field": label, "code": "REQUIRED_FIELD", "message": f"{label}不能为空"})
                elif len(normalized[field]) > max_length:
                    errors.append({"row": row_number, "field": label, "code": "FIELD_TOO_LONG", "message": f"{label}长度超出限制"})
            if normalized["imageUrl"]:
                parsed_url = urlparse(normalized["imageUrl"])
                if parsed_url.scheme not in {"https", "http"} or not parsed_url.netloc:
                    errors.append({"row": row_number, "field": "店铺图片", "code": "INVALID_URL", "message": "店铺图片必须是 HTTP(S) URL"})
            parsed.append(normalized)
        if errors:
            raise ApiError(422, "IMPORT_VALIDATION_FAILED", "导入文件包含错误，未写入任何店铺", details=errors)
        return parsed
    if headers != HEADERS:
        raise ApiError(
            422,
            "IMPORT_VALIDATION_FAILED",
            "表头必须为 storeCode,schoolCode,areaCode,name,category,address,imageUrl,status",
        )
    data_rows = [row for row in rows[1:] if any(_cell(value) for value in row)]
    if not data_rows:
        raise ApiError(422, "IMPORT_VALIDATION_FAILED", "导入文件没有数据行")
    if len(data_rows) > 1000:
        raise ApiError(422, "IMPORT_VALIDATION_FAILED", "数据行不能超过 1000 行")
    errors: list[dict] = []
    parsed: list[dict[str, str]] = []
    seen: set[str] = set()
    for row_number, row in enumerate(data_rows, start=2):
        values = [_cell(value) for value in row]
        values += [""] * (len(HEADERS) - len(values))
        item = dict(zip(HEADERS, values[: len(HEADERS)]))
        item["storeCode"] = item["storeCode"].lower()
        item["schoolCode"] = item["schoolCode"].lower()
        item["areaCode"] = item["areaCode"].lower()
        if item["storeCode"] in seen:
            errors.append({"row": row_number, "field": "storeCode", "code": "DUPLICATE_STORE_CODE_IN_FILE", "message": "店铺编码在文件中重复"})
        seen.add(item["storeCode"])
        for field, max_length in [
            ("storeCode", 100),
            ("schoolCode", 64),
            ("areaCode", 64),
            ("name", 120),
            ("category", 50),
            ("address", 255),
        ]:
            if not item[field]:
                errors.append({"row": row_number, "field": field, "code": "REQUIRED_FIELD", "message": "字段不能为空"})
            elif len(item[field]) > max_length:
                errors.append({"row": row_number, "field": field, "code": "FIELD_TOO_LONG", "message": "字段长度超出限制"})
        if item["storeCode"] and not re.fullmatch(r"[a-z0-9_-]{2,100}", item["storeCode"]):
            errors.append({"row": row_number, "field": "storeCode", "code": "INVALID_STORE_CODE", "message": "店铺编码只能包含小写字母、数字、- 和 _"})
        for field, label in [("schoolCode", "学校编码"), ("areaCode", "区域编码")]:
            if item[field] and not re.fullmatch(r"[a-z0-9_-]{2,64}", item[field]):
                errors.append(
                    {
                        "row": row_number,
                        "field": field,
                        "code": "INVALID_CODE",
                        "message": f"{label}只能包含小写字母、数字、- 和 _",
                    }
                )
        if item["status"] not in VALID_STATUSES:
            errors.append({"row": row_number, "field": "status", "code": "INVALID_ENUM_VALUE", "message": "状态必须为 active、hidden 或 closed"})
        if item["imageUrl"]:
            parsed_url = urlparse(item["imageUrl"])
            if parsed_url.scheme != "https" and parsed_url.scheme != "http":
                errors.append({"row": row_number, "field": "imageUrl", "code": "INVALID_URL", "message": "图片地址必须是 HTTP(S) URL"})
        parsed.append(item)
    if errors:
        raise ApiError(422, "IMPORT_VALIDATION_FAILED", "导入文件包含错误，未写入任何店铺", details=errors)
    return parsed


async def import_stores(
    session: AsyncSession,
    storage: MinioStorage,
    settings: Settings,
    content: bytes,
    filename: str,
    *,
    allowed_school_ids: set[int] | None = None,
    target_school_id: int | None = None,
) -> dict:
    rows = await parse_rows(content, filename, settings, target_school_id=target_school_id)
    if target_school_id is not None:
        if allowed_school_ids is not None and target_school_id not in allowed_school_ids:
            raise ApiError(403, "FORBIDDEN", "导入目标学校不在当前管理员权限范围内")
        school = await session.get(School, target_school_id)
        if school is None or school.status != "active":
            raise ApiError(422, "SCHOOL_NOT_FOUND", "目标学校不存在或未启用")
        area_names = {row["areaName"].strip().lower() for row in rows}
        area_rows = list((await session.scalars(select(SchoolArea).where(SchoolArea.school_id == target_school_id, SchoolArea.status == "active"))).all())
        area_map = {area.name.strip().lower(): area for area in area_rows}
        # 同时允许模板填写区域编码，方便已有数据直接复制。
        area_map.update({area.area_code.strip().lower(): area for area in area_rows})
        unknown = sorted(name for name in area_names if name not in area_map)
        if unknown:
            raise ApiError(422, "IMPORT_VALIDATION_FAILED", "模板中的食堂不属于目标学校或未启用", details=[{"field": "食堂", "code": "UNKNOWN_SCHOOL_AREA", "message": name} for name in unknown])
        existing = {}
        results = []
        try:
            for row_index, row in enumerate(rows, start=2):
                area = area_map[row["areaName"].strip().lower()]
                digest = hashlib.sha1(f"{target_school_id}|{area.id}|{row['name'].strip().lower()}|{row['address'].strip().lower()}".encode("utf-8")).hexdigest()[:12]
                store_code = f"imp-{target_school_id}-{area.id}-{digest}"
                store = await session.scalar(select(Store).where(Store.store_code == store_code).with_for_update())
                action = "updated" if store else "created"
                if store is None:
                    store = Store(store_code=store_code, school_id=target_school_id, area_id=area.id, name=row["name"], address=row["address"], status="active")
                    session.add(store)
                    await session.flush()
                else:
                    store.name = row["name"]
                    store.address = row["address"]
                    store.status = "active"
                    store.version += 1
                categories = await ensure_categories(session, "未分类")
                await replace_categories(session, store.id, categories)
                await attach_image(session, store, row["imageUrl"], storage)
                results.append({"row": row_index, "store_code": store.store_code, "store_id": str(store.id), "action": action})
            await session.commit()
        except Exception:
            await session.rollback()
            raise
        return {"total_rows": len(rows), "created_count": sum(item["action"] == "created" for item in results), "updated_count": sum(item["action"] == "updated" for item in results), "items": results}
    codes = [row["storeCode"] for row in rows]
    requested_areas = {(row["schoolCode"], row["areaCode"]) for row in rows}
    school_codes = {school_code for school_code, _ in requested_areas}
    area_codes = {area_code for _, area_code in requested_areas}
    area_rows = (
        await session.execute(
            select(School.school_code, SchoolArea.area_code, School.id, SchoolArea.id)
            .join(SchoolArea, SchoolArea.school_id == School.id)
            .where(School.school_code.in_(school_codes), SchoolArea.area_code.in_(area_codes))
        )
    ).all()
    area_map = {
        (school_code, area_code): (school_id, area_id)
        for school_code, area_code, school_id, area_id in area_rows
    }
    unknown_areas = sorted(requested_areas - set(area_map))
    if unknown_areas:
        details = [
            {
                "field": "areaCode",
                "code": "UNKNOWN_SCHOOL_AREA",
                "message": f"学校或区域不存在：{school_code}/{area_code}",
            }
            for school_code, area_code in unknown_areas
        ]
        raise ApiError(422, "IMPORT_VALIDATION_FAILED", "导入文件引用了不存在的学校或区域", details=details)
    if allowed_school_ids is not None:
        forbidden = sorted(
            f"{school_code}/{area_code}"
            for (school_code, area_code), (school_id, _) in area_map.items()
            if school_id not in allowed_school_ids
        )
        if forbidden:
            raise ApiError(
                403,
                "FORBIDDEN",
                "导入文件包含无权管理的学校",
                details=[{"field": "schoolCode", "code": "SCHOOL_SCOPE_FORBIDDEN", "message": value} for value in forbidden],
            )

    existing = {
        store.store_code: store
        for store in (await session.scalars(select(Store).where(Store.store_code.in_(codes)))).all()
    }
    results = []
    try:
        for row_index, row in enumerate(rows, start=2):
            store = existing.get(row["storeCode"])
            action = "updated" if store else "created"
            school_id, area_id = area_map[(row["schoolCode"], row["areaCode"])]
            if store is None:
                store = Store(
                    store_code=row["storeCode"],
                    school_id=school_id,
                    area_id=area_id,
                    name=row["name"],
                    address=row["address"],
                    status=row["status"],
                )
                session.add(store)
                await session.flush()
                existing[store.store_code] = store
            else:
                store.school_id = school_id
                store.area_id = area_id
                store.name = row["name"]
                store.address = row["address"]
                store.status = row["status"]
                store.version += 1
            categories = await ensure_categories(session, row["category"])
            await replace_categories(session, store.id, categories)
            if row["imageUrl"]:
                await attach_image(session, store, row["imageUrl"], storage)
            elif store.status == "active" and not store.images:
                raise ApiError(422, "IMPORT_VALIDATION_FAILED", "已上架店铺必须提供图片", field="imageUrl")
            results.append({"row": row_index, "store_code": store.store_code, "store_id": str(store.id), "action": action})
        await session.commit()
    except Exception:
        await session.rollback()
        raise
    return {"total_rows": len(rows), "created_count": sum(item["action"] == "created" for item in results), "updated_count": sum(item["action"] == "updated" for item in results), "items": results}
